import os
import time
import psutil
import psycopg2
import json
import requests
from psycopg2 import OperationalError
from openpyxl import Workbook, load_workbook
from datetime import datetime
from pathlib import Path

# ============================================================
# Configuration
# ============================================================

# Database Configuration
HOST = "localhost"
DATABASE = "linux_monitor"
USER = "postgres"
PASSWORD = "YOUR_PASSWORD_OF_DBMS"
PORT = "5432"

# Monitoring Configuration
INTERVAL = 60  # seconds

# Alert Thresholds (Percentage)
CPU_THRESHOLD = 85.0
MEMORY_THRESHOLD = 90.0
DISK_THRESHOLD = 90.0

# Optional Webhook URL (Slack, Discord, Teams, or custom API endpoint)
# Leave empty string "" to disable webhook alerts
WEBHOOK_URL = ""

# Alert Cooldown (seconds) - prevents spamming alerts continuously
ALERT_COOLDOWN = 900  # 15 minutes
last_alert_time = 0

# ============================================================
# Get Disk Path
# ============================================================

DISK_PATH = "C:\\" if os.name == "nt" else "/"

# ============================================================
# Dynamic Excel Filename (Daily Rotation)
# ============================================================

def get_excel_filename():
    """Generates a separate Excel file name for each day (e.g., Server_Monitoring_2026-08-08.xlsx)."""
    today_str = datetime.now().strftime("%Y-%m-%d")
    return f"Server_Monitoring_{today_str}.xlsx"

# ============================================================
# Alert System
# ============================================================

def check_and_alert(cpu, memory, disk, recorded_at):
    global last_alert_time

    alerts = []
    if cpu >= CPU_THRESHOLD:
        alerts.append(f"⚠️ HIGH CPU USAGE: {cpu:.2f}% (Threshold: {CPU_THRESHOLD}%)")
    if memory >= MEMORY_THRESHOLD:
        alerts.append(f"⚠️ HIGH MEMORY USAGE: {memory:.2f}% (Threshold: {MEMORY_THRESHOLD}%)")
    if disk >= DISK_THRESHOLD:
        alerts.append(f"⚠️ HIGH DISK USAGE: {disk:.2f}% (Threshold: {DISK_THRESHOLD}%)")

    if not alerts:
        return

    # Print local terminal alert
    timestamp_str = recorded_at.strftime("%Y-%m-%d %H:%M:%S")
    print(f"\n[ALERT - {timestamp_str}]")
    for alert in alerts:
        print(f"  - {alert}")
    print()

    # Dispatch Webhook if URL is configured and cooldown has passed
    current_time = time.time()
    if WEBHOOK_URL and (current_time - last_alert_time) >= ALERT_COOLDOWN:
        payload = {
            "text": f"*Server Alert [{timestamp_str}]*\n" + "\n".join(alerts)
        }
        try:
            response = requests.post(WEBHOOK_URL, json=payload, timeout=5)
            if response.status_code == 200:
                print("Webhook alert dispatched successfully.")
                last_alert_time = current_time
            else:
                print(f"Webhook failed with status code: {response.status_code}")
        except Exception as e:
            print(f"Failed to send webhook alert: {e}")

# ============================================================
# PostgreSQL Connection & Setup
# ============================================================

def get_connection():
    return psycopg2.connect(
        host=HOST,
        database=DATABASE,
        user=USER,
        password=PASSWORD,
        port=PORT,
        connect_timeout=5
    )

def create_table():
    conn = None
    cur = None
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("""
            CREATE TABLE IF NOT EXISTS server_metrics (
                id SERIAL PRIMARY KEY,
                cpu_usage DECIMAL(5,2) NOT NULL,
                memory_usage DECIMAL(5,2) NOT NULL,
                disk_usage DECIMAL(5,2) NOT NULL,
                recorded_at TIMESTAMP NOT NULL
            )
        """)
        conn.commit()
        print("PostgreSQL table ready.")
    except OperationalError as e:
        print("Database connection error:", e)
    except Exception as e:
        print("Database error:", e)
    finally:
        if cur: cur.close()
        if conn: conn.close()

# ============================================================
# Metrics Collection & DB Storage
# ============================================================

def get_metrics():
    try:
        cpu = psutil.cpu_percent(interval=1)
        memory = psutil.virtual_memory().percent
        disk = psutil.disk_usage(DISK_PATH).percent
        return round(cpu, 2), round(memory, 2), round(disk, 2)
    except Exception as e:
        print("Error reading system metrics:", e)
        return None

def save_to_database(cpu, memory, disk, recorded_at):
    conn = None
    cur = None
    try:
        conn = get_connection()
        cur = conn.cursor()
        cur.execute("""
            INSERT INTO server_metrics (cpu_usage, memory_usage, disk_usage, recorded_at)
            VALUES (%s, %s, %s, %s)
        """, (cpu, memory, disk, recorded_at))
        conn.commit()
        return True
    except Exception as e:
        print("PostgreSQL error:", e)
        return False
    finally:
        if cur: cur.close()
        if conn: conn.close()

# ============================================================
# Daily Excel Storage
# ============================================================

def create_excel_file(excel_file):
    file_path = Path(excel_file)
    if not file_path.exists():
        wb = Workbook()
        ws = wb.active
        ws.title = "Server Monitoring"
        ws.append(["ID", "Date", "Time", "CPU %", "Memory %", "Disk %", "Timestamp"])

        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 12
        ws.column_dimensions["E"].width = 12
        ws.column_dimensions["F"].width = 12
        ws.column_dimensions["G"].width = 25

        wb.save(excel_file)
        wb.close()
        print(f"Created new daily log file: {excel_file}")

def save_to_excel(cpu, memory, disk, recorded_at):
    excel_file = get_excel_filename()
    try:
        create_excel_file(excel_file)
        wb = load_workbook(excel_file)
        ws = wb["Server Monitoring"] if "Server Monitoring" in wb.sheetnames else wb.create_sheet("Server Monitoring")

        next_id = ws.max_row
        date_value = recorded_at.strftime("%Y-%m-%d")
        time_value = recorded_at.strftime("%H:%M:%S")
        timestamp_value = recorded_at.strftime("%Y-%m-%d %H:%M:%S")

        ws.append([
            next_id,
            date_value,
            time_value,
            cpu,
            memory,
            disk,
            timestamp_value
        ])

        wb.save(excel_file)
        wb.close()
        return True
    except PermissionError:
        print(f"\nERROR: Excel file {excel_file} is open. Please close it.\n")
        return False
    except Exception as e:
        print("Excel error:", e)
        return False

# ============================================================
# Display & Main Loop
# ============================================================

def display_metrics(cpu, memory, disk, recorded_at, db_status, excel_status):
    print(
        f"{recorded_at.strftime('%Y-%m-%d %H:%M:%S')} | "
        f"CPU: {cpu:6.2f}% | "
        f"Memory: {memory:6.2f}% | "
        f"Disk: {disk:6.2f}% | "
        f"DB: {db_status} | "
        f"Excel: {excel_status}"
    )

def main():
    print("=" * 85)
    print("              SERVER MONITORING SYSTEM")
    print("=" * 85)
    print(f"Operating System : {os.name}")
    print(f"Disk Monitoring  : {DISK_PATH}")
    print(f"Database         : {DATABASE}")
    print(f"Excel Mode       : Daily File Rotation")
    print(f"Interval         : {INTERVAL} seconds")
    print("=" * 85)

    create_table()

    print("\nMonitoring started... Press Ctrl + C to stop.\n")
    record_number = 0

    try:
        while True:
            recorded_at = datetime.now()
            metrics = get_metrics()

            if metrics is None:
                print("Unable to collect system metrics.")
                time.sleep(INTERVAL)
                continue

            cpu, memory, disk = metrics

            # Check threshold limits
            check_and_alert(cpu, memory, disk, recorded_at)

            # Persist data
            db_saved = save_to_database(cpu, memory, disk, recorded_at)
            excel_saved = save_to_excel(cpu, memory, disk, recorded_at)

            record_number += 1
            db_status = "OK" if db_saved else "FAIL"
            excel_status = "OK" if excel_saved else "FAIL"

            display_metrics(cpu, memory, disk, recorded_at, db_status, excel_status)
            time.sleep(INTERVAL)

    except KeyboardInterrupt:
        print("\n" + "=" * 85)
        print("Monitoring stopped by user.")
        print(f"Total records collected this session: {record_number}")
        print("=" * 85)

if __name__ == "__main__":
    main()